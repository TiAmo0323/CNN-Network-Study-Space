from pathlib import Path

import torch
import torch.nn as nn
import torch.optim as optim
import torchvision
import torchvision.transforms as transforms
from openpyxl import Workbook, load_workbook

from model import LeNet


EXCEL_HEADERS = [
    '实验类别', '实验名称', '计划训练轮数', '优化器', '优化器参数',
    '学习率', 'epoch', 'step', 'train_loss', 'test_accuracy'
]

# 每组实验只改变一个主要变量，未改变的参数沿用官方Demo设置。
EXPERIMENTS = [
    # 增加训练轮数：保持Adam和学习率0.001不变
    {
        'category': '训练轮数对照',
        'name': 'Adam_lr0.001_10轮',
        'epochs': 10,
        'optimizer': 'Adam',
        'optimizer_params': '默认参数',
        'learning_rate': 0.001,
        'model_file': 'Lenet_epochs_10.pth',
    },
    {
        'category': '训练轮数对照',
        'name': 'Adam_lr0.001_15轮',
        'epochs': 15,
        'optimizer': 'Adam',
        'optimizer_params': '默认参数',
        'learning_rate': 0.001,
        'model_file': 'Lenet_epochs_15.pth',
    },
    {
        'category': '训练轮数对照',
        'name': 'Adam_lr0.001_20轮',
        'epochs': 20,
        'optimizer': 'Adam',
        'optimizer_params': '默认参数',
        'learning_rate': 0.001,
        'model_file': 'Lenet_epochs_20.pth',
    },

    # 修改学习率：保持Adam和训练5轮不变
    {
        'category': '学习率对照',
        'name': 'Adam_lr0.002_5轮',
        'epochs': 5,
        'optimizer': 'Adam',
        'optimizer_params': '默认参数',
        'learning_rate': 0.002,
        'model_file': 'Lenet_lr_0.002.pth',
    },
    {
        'category': '学习率对照',
        'name': 'Adam_lr0.003_5轮',
        'epochs': 5,
        'optimizer': 'Adam',
        'optimizer_params': '默认参数',
        'learning_rate': 0.003,
        'model_file': 'Lenet_lr_0.003.pth',
    },
    {
        'category': '学习率对照',
        'name': 'Adam_lr0.005_5轮',
        'epochs': 5,
        'optimizer': 'Adam',
        'optimizer_params': '默认参数',
        'learning_rate': 0.005,
        'model_file': 'Lenet_lr_0.005.pth',
    },

    # 修改优化器：保持学习率0.001和训练5轮不变
    {
        'category': '优化器对照',
        'name': 'SGD_lr0.001_5轮',
        'epochs': 5,
        'optimizer': 'SGD',
        'optimizer_params': '默认参数',
        'learning_rate': 0.001,
        'model_file': 'Lenet_optimizer_SGD.pth',
    },
    {
        'category': '优化器对照',
        'name': 'RMSprop_lr0.001_5轮',
        'epochs': 5,
        'optimizer': 'RMSprop',
        'optimizer_params': '默认参数',
        'learning_rate': 0.001,
        'model_file': 'Lenet_optimizer_RMSprop.pth',
    },
    {
        'category': '优化器对照',
        'name': 'Adagrad_lr0.001_5轮',
        'epochs': 5,
        'optimizer': 'Adagrad',
        'optimizer_params': '默认参数',
        'learning_rate': 0.001,
        'model_file': 'Lenet_optimizer_Adagrad.pth',
    },
]


def prepare_workbook(excel_path):
    """打开结果表，并兼容此前只有4列的官方Demo结果。"""
    if excel_path.exists():
        workbook = load_workbook(excel_path)
        worksheet = (
            workbook['训练数据']
            if '训练数据' in workbook.sheetnames
            else workbook.active
        )
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '训练数据'

    current_headers = [
        worksheet.cell(row=1, column=column).value
        for column in range(1, worksheet.max_column + 1)
    ]

    legacy_headers = ['epoch', 'step', 'train_loss', 'test_accuracy']
    if current_headers == legacy_headers:
        # 原4列数据右移，在前面补充官方Demo的实验说明。
        worksheet.insert_cols(1, amount=6)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=1, value='官方基线')
            worksheet.cell(row=row, column=2, value='官方Demo_Adam_lr0.001_5轮')
            worksheet.cell(row=row, column=3, value=5)
            worksheet.cell(row=row, column=4, value='Adam')
            worksheet.cell(row=row, column=5, value='默认参数')
            worksheet.cell(row=row, column=6, value=0.001)

    for column, header in enumerate(EXCEL_HEADERS, start=1):
        worksheet.cell(row=1, column=column, value=header)

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = 'A1:J1'
    workbook.save(excel_path)
    return workbook, worksheet


def create_optimizer(optimizer_name, parameters, learning_rate):
    if optimizer_name == 'Adam':
        return optim.Adam(parameters, lr=learning_rate)
    if optimizer_name == 'SGD':
        return optim.SGD(parameters, lr=learning_rate)
    if optimizer_name == 'RMSprop':
        return optim.RMSprop(parameters, lr=learning_rate)
    if optimizer_name == 'Adagrad':
        return optim.Adagrad(parameters, lr=learning_rate)
    raise ValueError('不支持的优化器：{}'.format(optimizer_name))


def remove_previous_results(worksheet, experiment_name):
    """重跑某组实验时先移除该组旧记录，避免Excel中出现重复数据。"""
    for row in range(worksheet.max_row, 1, -1):
        if worksheet.cell(row=row, column=2).value == experiment_name:
            worksheet.delete_rows(row)


def run_experiment(
        experiment, train_loader, val_image, val_label,
        loss_function, workbook, worksheet, excel_path, models_dir):
    # 固定随机种子，使不同实验使用相同的初始权重和数据打乱序列。
    torch.manual_seed(42)
    net = LeNet()
    optimizer = create_optimizer(
        experiment['optimizer'],
        net.parameters(),
        experiment['learning_rate'],
    )

    remove_previous_results(worksheet, experiment['name'])
    workbook.save(excel_path)

    print('\n开始实验：{}（类别：{}）'.format(
        experiment['name'], experiment['category']
    ))

    for epoch in range(experiment['epochs']):
        net.train()
        running_loss = 0.0

        for step, data in enumerate(train_loader, start=0):
            inputs, labels = data

            optimizer.zero_grad()
            outputs = net(inputs)
            loss = loss_function(outputs, labels)
            loss.backward()
            optimizer.step()

            running_loss += loss.item()
            if step % 500 == 499:
                net.eval()
                with torch.no_grad():
                    outputs = net(val_image)
                    predict_y = torch.max(outputs, dim=1)[1]
                    accuracy = (
                        torch.eq(predict_y, val_label).sum().item()
                        / val_label.size(0)
                    )

                train_loss = running_loss / 500
                print(
                    '[{}] [{:d}, {:5d}] train_loss: {:.3f}  '
                    'test_accuracy: {:.3f}'.format(
                        experiment['name'], epoch + 1, step + 1,
                        train_loss, accuracy
                    )
                )

                worksheet.append([
                    experiment['category'],
                    experiment['name'],
                    experiment['epochs'],
                    experiment['optimizer'],
                    experiment['optimizer_params'],
                    experiment['learning_rate'],
                    epoch + 1,
                    step + 1,
                    train_loss,
                    accuracy,
                ])
                # 每次打印后立即保存，训练意外中断时仍能保留已有结果。
                workbook.save(excel_path)
                running_loss = 0.0
                net.train()

    model_path = models_dir / experiment['model_file']
    torch.save(net.state_dict(), model_path)
    print('实验完成，模型保存至：{}'.format(model_path))


def main():
    script_dir = Path(__file__).resolve().parent
    results_dir = script_dir / 'results'
    models_dir = results_dir / 'models'
    excel_path = results_dir / 'hyperparameter_results.xlsx'
    try:
        results_dir.mkdir(parents=True, exist_ok=True)
        models_dir.mkdir(parents=True, exist_ok=True)
        workbook, worksheet = prepare_workbook(excel_path)
    except PermissionError:
        print('无法写入Excel：{}'.format(excel_path))
        print('请先关闭正在使用该文件的Excel或WPS，然后重新运行train.py。')
        return

    transform = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5)),
    ])

    # 复用LeNet_experiment根目录中的CIFAR-10数据集。
    data_path = script_dir.parent / 'data'
    train_set = torchvision.datasets.CIFAR10(
        root=str(data_path), train=True, download=True, transform=transform
    )
    train_loader = torch.utils.data.DataLoader(
        train_set, batch_size=36, shuffle=True, num_workers=0
    )

    val_set = torchvision.datasets.CIFAR10(
        root=str(data_path), train=False, download=True, transform=transform
    )
    val_loader = torch.utils.data.DataLoader(
        val_set, batch_size=5000, shuffle=False, num_workers=0
    )
    val_image, val_label = next(iter(val_loader))

    loss_function = nn.CrossEntropyLoss()

    try:
        for experiment in EXPERIMENTS:
            run_experiment(
                experiment,
                train_loader,
                val_image,
                val_label,
                loss_function,
                workbook,
                worksheet,
                excel_path,
                models_dir,
            )
    finally:
        workbook.save(excel_path)
        workbook.close()

    print('\n全部对照实验完成')
    print('实验数据保存至：{}'.format(excel_path))


if __name__ == '__main__':
    main()
