"""绘制LeNet官方Demo及三类对照实验结果。"""

import math
from collections import defaultdict
from pathlib import Path

import matplotlib

matplotlib.use('Agg')

import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator, PercentFormatter
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / 'results' / 'hyperparameter_results.xlsx'
OUTPUT_DIR = SCRIPT_DIR / 'results' / 'figures'

TRAIN_SET_SIZE = 50000
BATCH_SIZE = 36
BATCHES_PER_EPOCH = math.ceil(TRAIN_SET_SIZE / BATCH_SIZE)

CATEGORY_SETTINGS = {
    '训练轮数对照': {
        'title': '不同训练轮数对照（Adam，学习率=0.001）',
        'filename': '训练轮数对照.png',
    },
    '学习率对照': {
        'title': '不同学习率对照（Adam，训练5轮）',
        'filename': '学习率对照.png',
    },
    '优化器对照': {
        'title': '不同优化器对照（学习率=0.001，训练5轮）',
        'filename': '优化器对照.png',
    },
}

REQUIRED_HEADERS = {
    '实验类别', '实验名称', '计划训练轮数', '优化器', '学习率',
    'epoch', 'step', 'train_loss', 'test_accuracy'
}


def configure_chinese_font():
    """优先使用Windows常见中文字体，并修复负号显示。"""
    plt.rcParams['font.sans-serif'] = [
        'Microsoft YaHei', 'SimHei', 'Noto Sans CJK SC', 'DejaVu Sans'
    ]
    plt.rcParams['axes.unicode_minus'] = False


def load_records(excel_path):
    if not excel_path.exists():
        raise FileNotFoundError('未找到实验数据：{}'.format(excel_path))

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        worksheet = (
            workbook['训练数据']
            if '训练数据' in workbook.sheetnames
            else workbook.active
        )
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows))

        missing_headers = REQUIRED_HEADERS.difference(headers)
        if missing_headers:
            raise ValueError(
                'Excel缺少必要字段：{}'.format('、'.join(sorted(missing_headers)))
            )

        records = []
        for values in rows:
            record = dict(zip(headers, values))
            if not record.get('实验名称'):
                continue
            if any(record.get(key) is None for key in (
                    'epoch', 'step', 'train_loss', 'test_accuracy')):
                continue
            records.append(record)
    finally:
        workbook.close()

    if not records:
        raise ValueError('Excel中没有可绘制的实验记录。')
    return records


def experiment_label(records):
    first = records[0]
    if first['实验类别'] == '官方基线':
        return '官方基线：Adam，lr=0.001，5轮'
    return str(first['实验名称'])


def sort_experiment_names(grouped_records):
    """官方基线在前，其余实验按训练轮数、学习率和名称排序。"""
    def sort_key(name):
        record = grouped_records[name][0]
        is_not_baseline = record['实验类别'] != '官方基线'
        return (
            is_not_baseline,
            int(record['计划训练轮数']),
            float(record['学习率']),
            str(record['优化器']),
            name,
        )

    return sorted(grouped_records, key=sort_key)


def plot_category(all_records, category, settings, output_dir):
    selected_records = [
        record for record in all_records
        if record['实验类别'] in ('官方基线', category)
    ]
    category_records = [
        record for record in selected_records
        if record['实验类别'] == category
    ]
    if not category_records:
        print('跳过“{}”：Excel中尚无该类别的数据。'.format(category))
        return None

    grouped_records = defaultdict(list)
    for record in selected_records:
        grouped_records[str(record['实验名称'])].append(record)

    configure_chinese_font()
    figure, axes = plt.subplots(1, 2, figsize=(14, 5.8))
    colors = plt.get_cmap('tab10').colors
    line_styles = ['--', '-', '-.', ':']

    all_accuracies = []
    experiment_names = sort_experiment_names(grouped_records)
    for index, name in enumerate(experiment_names):
        records = sorted(
            grouped_records[name],
            key=lambda item: (int(item['epoch']), int(item['step']))
        )
        progress = [
            int(record['epoch']) - 1
            + int(record['step']) / BATCHES_PER_EPOCH
            for record in records
        ]
        losses = [float(record['train_loss']) for record in records]
        accuracies = [float(record['test_accuracy']) * 100 for record in records]
        all_accuracies.extend(accuracies)

        is_baseline = records[0]['实验类别'] == '官方基线'
        color = '#202020' if is_baseline else colors[(index - 1) % len(colors)]
        line_style = '--' if is_baseline else line_styles[(index - 1) % len(line_styles)]
        line_width = 2.4 if is_baseline else 1.9
        label = experiment_label(records)

        axes[0].plot(
            progress, losses, label=label, color=color,
            linestyle=line_style, linewidth=line_width,
            marker='o', markersize=3.5
        )
        axes[1].plot(
            progress, accuracies, label=label, color=color,
            linestyle=line_style, linewidth=line_width,
            marker='o', markersize=3.5
        )

    axes[0].set_title('训练损失变化')
    axes[0].set_xlabel('累计训练进度（epoch）')
    axes[0].set_ylabel('train_loss')

    axes[1].set_title('测试准确率变化')
    axes[1].set_xlabel('累计训练进度（epoch）')
    axes[1].set_ylabel('test_accuracy')
    axes[1].yaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=0))
    accuracy_min = max(0, min(all_accuracies) - 5)
    accuracy_max = min(100, max(all_accuracies) + 5)
    axes[1].set_ylim(accuracy_min, accuracy_max)

    for axis in axes:
        axis.grid(True, linestyle='--', alpha=0.35)
        axis.xaxis.set_major_locator(MaxNLocator(nbins=10))

    handles, labels = axes[0].get_legend_handles_labels()
    figure.legend(
        handles, labels, loc='lower center', ncol=2,
        frameon=False, bbox_to_anchor=(0.5, 0.01)
    )
    figure.suptitle(settings['title'], fontsize=15, fontweight='bold')
    figure.tight_layout(rect=(0, 0.15, 1, 0.94))

    output_path = output_dir / settings['filename']
    figure.savefig(output_path, dpi=200, bbox_inches='tight')
    plt.close(figure)
    return output_path


def main():
    records = load_records(EXCEL_PATH)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    generated_files = []
    for category, settings in CATEGORY_SETTINGS.items():
        output_path = plot_category(
            records, category, settings, OUTPUT_DIR
        )
        if output_path is not None:
            generated_files.append(output_path)

    if generated_files:
        print('可视化完成，共生成{}张图片：'.format(len(generated_files)))
        for file_path in generated_files:
            print(file_path)
    else:
        print('没有生成图片，请先完成对照实验并检查Excel中的实验类别。')


if __name__ == '__main__':
    main()
