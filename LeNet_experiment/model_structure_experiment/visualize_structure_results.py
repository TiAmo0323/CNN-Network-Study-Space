"""将模型结构消融实验Excel结果绘制为分类对照图。"""

import argparse
import math
from collections import defaultdict
from pathlib import Path

import matplotlib

matplotlib.use('Agg')

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MaxNLocator, PercentFormatter
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = SCRIPT_DIR / 'results' / 'model_structure_results.xlsx'
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / 'results' / 'figures'

TRAIN_SET_SIZE = 50000
BATCH_SIZE = 36
BATCHES_PER_EPOCH = math.ceil(TRAIN_SET_SIZE / BATCH_SIZE)

CATEGORY_SETTINGS = {
    '卷积通道消融': ('卷积通道数量消融', '02_卷积通道消融.png'),
    '网络深度消融': ('网络深度消融', '03_网络深度消融.png'),
    '卷积核大小消融': ('卷积核大小消融', '04_卷积核大小消融.png'),
    '池化方式消融': ('Pooling方式消融', '05_池化方式消融.png'),
    '激活函数消融': ('激活函数消融', '06_激活函数消融.png'),
    'Batch Normalization消融': (
        'Batch Normalization消融', '07_BatchNorm消融.png'
    ),
    'Dropout消融': ('Dropout正则化消融', '08_Dropout消融.png'),
}

DISPLAY_NAMES = {
    'baseline_current': '基线 16→32',
    'channel_reduced_4_8': '通道 4→8',
    'channel_classic_6_16': '通道 6→16',
    'depth_add_conv': '增加Conv3',
    'kernel_size_3': 'kernel=3',
    'kernel_size_7': 'kernel=7',
    'pooling_avg': 'AvgPool',
    'pooling_none': '无Pooling',
    'activation_sigmoid': 'Sigmoid',
    'activation_tanh': 'Tanh',
    'activation_leaky_relu': 'LeakyReLU',
    'batch_norm': '加入BatchNorm',
    'dropout_0_5': 'Dropout=0.5',
}

LOG_REQUIRED_HEADERS = {
    'Experiment', 'Category', 'Epoch', 'Step',
    'Train Loss', 'Test Accuracy'
}
SUMMARY_REQUIRED_HEADERS = {
    'Experiment', 'Parameters', 'Accuracy', 'Loss', 'Category',
    'Best Accuracy', 'Min Loss', 'Training Time (s)'
}


def configure_chinese_font():
    plt.rcParams['font.sans-serif'] = [
        'Microsoft YaHei', 'SimHei', 'Noto Sans CJK SC', 'DejaVu Sans'
    ]
    plt.rcParams['axes.unicode_minus'] = False


def read_worksheet(worksheet, required_headers):
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    missing_headers = required_headers.difference(headers)
    if missing_headers:
        raise ValueError(
            '{}缺少字段：{}'.format(
                worksheet.title, '、'.join(sorted(missing_headers))
            )
        )

    records = []
    for values in rows:
        record = dict(zip(headers, values))
        if record.get('Experiment'):
            records.append(record)
    return records


def load_results(excel_path):
    if not excel_path.exists():
        raise FileNotFoundError('未找到结构消融结果表：{}'.format(excel_path))

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        for sheet_name in ('training_log', 'summary'):
            if sheet_name not in workbook.sheetnames:
                raise ValueError('Excel缺少工作表：{}'.format(sheet_name))

        log_records = read_worksheet(
            workbook['training_log'], LOG_REQUIRED_HEADERS
        )
        summary_records = read_worksheet(
            workbook['summary'], SUMMARY_REQUIRED_HEADERS
        )
    finally:
        workbook.close()

    if not log_records or not summary_records:
        raise ValueError('结果表中没有可绘制的数据。')
    return log_records, summary_records


def display_name(experiment_name):
    return DISPLAY_NAMES.get(experiment_name, experiment_name)


def add_value_labels(axis, bars, formatter, fontsize=8):
    for bar in bars:
        height = bar.get_height()
        axis.annotate(
            formatter(height),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 3), textcoords='offset points',
            ha='center', va='bottom', fontsize=fontsize
        )


def plot_overview(summary_records, output_dir):
    records = sorted(
        summary_records,
        key=lambda item: (
            item['Experiment'] != 'baseline_current',
            str(item['Category']),
            str(item['Experiment']),
        )
    )
    names = [display_name(record['Experiment']) for record in records]
    positions = np.arange(len(records))
    bar_width = 0.38

    final_accuracies = np.array([
        float(record['Accuracy']) * 100 for record in records
    ])
    best_accuracies = np.array([
        float(record['Best Accuracy']) * 100 for record in records
    ])
    final_losses = np.array([
        float(record['Loss']) for record in records
    ])
    min_losses = np.array([
        float(record['Min Loss']) for record in records
    ])
    parameters = np.array([
        int(record['Parameters']) for record in records
    ])
    training_times = np.array([
        float(record['Training Time (s)']) for record in records
    ])

    configure_chinese_font()
    figure, axes = plt.subplots(2, 2, figsize=(18, 12))
    figure.suptitle('LeNet模型结构消融实验总体对比', fontsize=17,
                    fontweight='bold')

    final_bars = axes[0, 0].bar(
        positions - bar_width / 2, final_accuracies,
        width=bar_width, label='最终准确率', color='#4C78A8'
    )
    best_bars = axes[0, 0].bar(
        positions + bar_width / 2, best_accuracies,
        width=bar_width, label='最佳准确率', color='#72B7B2'
    )
    axes[0, 0].set_title('测试准确率对比')
    axes[0, 0].set_ylabel('Accuracy')
    axes[0, 0].yaxis.set_major_formatter(
        PercentFormatter(xmax=100, decimals=0)
    )
    axes[0, 0].set_ylim(
        max(0, min(final_accuracies.min(), best_accuracies.min()) - 8),
        min(100, max(final_accuracies.max(), best_accuracies.max()) + 8),
    )
    axes[0, 0].legend()
    add_value_labels(axes[0, 0], final_bars, lambda value: '{:.1f}%'.format(value))
    add_value_labels(axes[0, 0], best_bars, lambda value: '{:.1f}%'.format(value))

    loss_bars = axes[0, 1].bar(
        positions - bar_width / 2, final_losses,
        width=bar_width, label='最终损失', color='#F58518'
    )
    min_loss_bars = axes[0, 1].bar(
        positions + bar_width / 2, min_losses,
        width=bar_width, label='最低损失', color='#FFBF79'
    )
    axes[0, 1].set_title('训练损失对比')
    axes[0, 1].set_ylabel('Train Loss')
    axes[0, 1].legend()
    add_value_labels(axes[0, 1], loss_bars, lambda value: '{:.2f}'.format(value))
    add_value_labels(
        axes[0, 1], min_loss_bars, lambda value: '{:.2f}'.format(value)
    )

    parameter_bars = axes[1, 0].bar(
        positions, parameters, color='#54A24B'
    )
    axes[1, 0].set_title('可训练参数量对比（对数坐标）')
    axes[1, 0].set_ylabel('Parameters')
    axes[1, 0].set_yscale('log')
    add_value_labels(
        axes[1, 0], parameter_bars,
        lambda value: '{:.0f}K'.format(value / 1000)
        if value < 1000000 else '{:.2f}M'.format(value / 1000000)
    )

    time_bars = axes[1, 1].bar(
        positions, training_times, color='#E45756'
    )
    axes[1, 1].set_title('训练时间对比')
    axes[1, 1].set_ylabel('Time (s)')
    add_value_labels(
        axes[1, 1], time_bars, lambda value: '{:.0f}s'.format(value)
    )

    for axis in axes.flat:
        axis.set_xticks(positions)
        axis.set_xticklabels(names, rotation=35, ha='right')
        axis.grid(axis='y', linestyle='--', alpha=0.3)

    figure.tight_layout(rect=(0, 0, 1, 0.96))
    output_path = output_dir / '01_总体指标对比.png'
    figure.savefig(output_path, dpi=200, bbox_inches='tight')
    plt.close(figure)
    return output_path


def plot_category(log_records, category, title, filename, output_dir):
    selected_records = [
        record for record in log_records
        if record['Category'] in ('基线', category)
    ]
    category_records = [
        record for record in selected_records
        if record['Category'] == category
    ]
    if not category_records:
        print('跳过“{}”：没有该类别的训练记录。'.format(category))
        return None

    grouped_records = defaultdict(list)
    for record in selected_records:
        grouped_records[record['Experiment']].append(record)

    experiment_names = sorted(
        grouped_records,
        key=lambda name: (name != 'baseline_current', name)
    )

    configure_chinese_font()
    figure, axes = plt.subplots(1, 2, figsize=(14, 5.8))
    colors = plt.get_cmap('tab10').colors
    line_styles = ['--', '-', '-.', ':']
    all_accuracies = []

    for index, experiment_name in enumerate(experiment_names):
        records = sorted(
            grouped_records[experiment_name],
            key=lambda item: (int(item['Epoch']), int(item['Step']))
        )
        progress = [
            int(record['Epoch']) - 1
            + int(record['Step']) / BATCHES_PER_EPOCH
            for record in records
        ]
        losses = [float(record['Train Loss']) for record in records]
        accuracies = [
            float(record['Test Accuracy']) * 100 for record in records
        ]
        all_accuracies.extend(accuracies)

        is_baseline = experiment_name == 'baseline_current'
        color = '#202020' if is_baseline else colors[(index - 1) % len(colors)]
        line_style = '--' if is_baseline else line_styles[(index - 1) % len(line_styles)]
        line_width = 2.4 if is_baseline else 1.9
        label = display_name(experiment_name)

        axes[0].plot(
            progress, losses, label=label, color=color,
            linestyle=line_style, linewidth=line_width,
            marker='o', markersize=3.5
        )
        axes[1].plot(
            progress, accuracies, label=label, color=color,
            linestyle=line_style, linewidth=line_width,
            marker='o', markersize=3.5
        )

    axes[0].set_title('训练损失变化')
    axes[0].set_xlabel('累计训练进度（epoch）')
    axes[0].set_ylabel('Train Loss')

    axes[1].set_title('测试准确率变化')
    axes[1].set_xlabel('累计训练进度（epoch）')
    axes[1].set_ylabel('Test Accuracy')
    axes[1].yaxis.set_major_formatter(
        PercentFormatter(xmax=100, decimals=0)
    )
    axes[1].set_ylim(
        max(0, min(all_accuracies) - 5),
        min(100, max(all_accuracies) + 5),
    )

    for axis in axes:
        axis.grid(True, linestyle='--', alpha=0.35)
        axis.xaxis.set_major_locator(MaxNLocator(nbins=10))

    handles, labels = axes[0].get_legend_handles_labels()
    figure.legend(
        handles, labels, loc='lower center', ncol=2,
        frameon=False, bbox_to_anchor=(0.5, 0.01)
    )
    figure.suptitle(title, fontsize=15, fontweight='bold')
    figure.tight_layout(rect=(0, 0.12, 1, 0.94))

    output_path = output_dir / filename
    figure.savefig(output_path, dpi=200, bbox_inches='tight')
    plt.close(figure)
    return output_path


def parse_args():
    parser = argparse.ArgumentParser(
        description='模型结构消融实验结果可视化'
    )
    parser.add_argument(
        '--excel', type=Path, default=DEFAULT_EXCEL_PATH,
        help='模型结构实验Excel路径'
    )
    parser.add_argument(
        '--output-dir', type=Path, default=DEFAULT_OUTPUT_DIR,
        help='图片输出目录'
    )
    return parser.parse_args()


def main():
    args = parse_args()
    log_records, summary_records = load_results(args.excel.resolve())
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    generated_files = [plot_overview(summary_records, output_dir)]
    for category, (title, filename) in CATEGORY_SETTINGS.items():
        output_path = plot_category(
            log_records, category, title, filename, output_dir
        )
        if output_path is not None:
            generated_files.append(output_path)

    print('可视化完成，共生成{}张图片：'.format(len(generated_files)))
    for output_path in generated_files:
        print(output_path)


if __name__ == '__main__':
    main()
