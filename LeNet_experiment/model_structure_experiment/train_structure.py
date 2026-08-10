"""运行LeNet模型结构消融实验，并保存明细、汇总和模型权重。"""

import argparse
import time
from pathlib import Path

import torch
import torch.nn as nn
import torch.optim as optim
import torchvision
import torchvision.transforms as transforms
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from model_structure import MODEL_CONFIGS, count_parameters, create_model


SCRIPT_DIR = Path(__file__).resolve().parent
PARENT_EXPERIMENT_DIR = SCRIPT_DIR.parent
DATA_DIR = PARENT_EXPERIMENT_DIR / 'data'
RESULTS_DIR = SCRIPT_DIR / 'results'
MODELS_DIR = RESULTS_DIR / 'models'
EXCEL_PATH = RESULTS_DIR / 'model_structure_results.xlsx'
SUMMARY_PATH = RESULTS_DIR / 'summary.md'

OPTIMIZER_NAME = 'Adam'
LEARNING_RATE = 0.001
EPOCHS = 10
BATCH_SIZE = 36
VALIDATION_BATCH_SIZE = 5000
RANDOM_SEED = 42
LOG_INTERVAL = 500

LOG_HEADERS = [
    'Experiment', 'Category', 'Modification', 'Optimizer', 'Learning Rate',
    'Planned Epochs', 'Epoch', 'Step', 'Train Loss', 'Test Accuracy',
    'Parameters', 'Elapsed Time (s)'
]

SUMMARY_HEADERS = [
    'Experiment', 'Modification', 'Parameters', 'Accuracy', 'Loss',
    'Conclusion', 'Category', 'Best Accuracy', 'Min Loss',
    'Training Time (s)', 'Optimizer', 'Learning Rate', 'Epochs', 'Model File'
]


def create_or_get_sheet(workbook, sheet_name, headers):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = '{}1:{}1'.format(
        worksheet.cell(row=1, column=1).column_letter,
        worksheet.cell(row=1, column=len(headers)).column_letter,
    )
    return worksheet


def prepare_workbook(excel_path):
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    if excel_path.exists():
        workbook = load_workbook(excel_path)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    log_sheet = create_or_get_sheet(
        workbook, 'training_log', LOG_HEADERS
    )
    summary_sheet = create_or_get_sheet(
        workbook, 'summary', SUMMARY_HEADERS
    )

    # Excel中以百分比显示准确率，底层仍保存0~1浮点数。
    for row in range(2, log_sheet.max_row + 1):
        log_sheet.cell(row=row, column=10).number_format = '0.00%'
    for row in range(2, summary_sheet.max_row + 1):
        summary_sheet.cell(row=row, column=4).number_format = '0.00%'
        summary_sheet.cell(row=row, column=8).number_format = '0.00%'

    workbook.save(excel_path)
    return workbook, log_sheet, summary_sheet


def remove_previous_results(worksheet, experiment_name):
    for row in range(worksheet.max_row, 1, -1):
        if worksheet.cell(row=row, column=1).value == experiment_name:
            worksheet.delete_rows(row)


def read_baseline_metrics(summary_sheet):
    for row in range(2, summary_sheet.max_row + 1):
        if summary_sheet.cell(row=row, column=1).value == 'baseline_current':
            return {
                'parameters': summary_sheet.cell(row=row, column=3).value,
                'accuracy': summary_sheet.cell(row=row, column=4).value,
                'loss': summary_sheet.cell(row=row, column=5).value,
            }
    return None


def build_conclusion(experiment_name, parameters, accuracy, baseline_metrics):
    if experiment_name == 'baseline_current':
        return '当前项目模型结构基线，用于其余单因素实验对照。'
    if baseline_metrics is None:
        return '本次未运行且结果表中不存在基线，暂无法计算相对变化。'

    accuracy_delta = (accuracy - baseline_metrics['accuracy']) * 100
    parameter_delta = parameters - baseline_metrics['parameters']
    accuracy_text = '提高' if accuracy_delta >= 0 else '降低'
    parameter_text = '增加' if parameter_delta >= 0 else '减少'
    return (
        '较基线准确率{}{:.2f}个百分点，参数量{}{:,}。'.format(
            accuracy_text, abs(accuracy_delta),
            parameter_text, abs(parameter_delta)
        )
    )


def append_log_row(
        worksheet, config_name, config, parameter_number,
        epoch, step, train_loss, accuracy, elapsed_time):
    worksheet.append([
        config_name,
        config['category'],
        config['modification'],
        OPTIMIZER_NAME,
        LEARNING_RATE,
        EPOCHS,
        epoch,
        step,
        train_loss,
        accuracy,
        parameter_number,
        elapsed_time,
    ])
    worksheet.cell(
        row=worksheet.max_row, column=10
    ).number_format = '0.00%'


def append_summary_row(
        worksheet, config_name, config, parameter_number,
        final_accuracy, final_loss, best_accuracy, min_loss,
        training_time, model_path, conclusion):
    worksheet.append([
        config_name,
        config['modification'],
        parameter_number,
        final_accuracy,
        final_loss,
        conclusion,
        config['category'],
        best_accuracy,
        min_loss,
        training_time,
        OPTIMIZER_NAME,
        LEARNING_RATE,
        EPOCHS,
        str(model_path),
    ])
    worksheet.cell(
        row=worksheet.max_row, column=4
    ).number_format = '0.00%'
    worksheet.cell(
        row=worksheet.max_row, column=8
    ).number_format = '0.00%'


def write_markdown_summary(summary_sheet, output_path):
    rows = []
    for row in summary_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)

    lines = [
        '# 模型结构消融实验汇总',
        '',
        '| Experiment | Modification | Parameters | Accuracy | Loss | Conclusion |',
        '| --- | --- | ---: | ---: | ---: | --- |',
    ]
    for row in rows:
        lines.append(
            '| {} | {} | {:,} | {:.2%} | {:.4f} | {} |'.format(
                row[0], row[1], int(row[2]), float(row[3]),
                float(row[4]), row[5]
            )
        )

    output_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')


def run_experiment(
        config_name, train_loader, val_image, val_label,
        loss_function, workbook, log_sheet, summary_sheet):
    config = MODEL_CONFIGS[config_name]
    torch.manual_seed(RANDOM_SEED)

    model = create_model(config_name)
    parameter_number = count_parameters(model)
    optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)

    remove_previous_results(log_sheet, config_name)
    remove_previous_results(summary_sheet, config_name)
    workbook.save(EXCEL_PATH)

    print('\n开始结构消融实验：{}'.format(config_name))
    print('修改：{}'.format(config['modification']))
    print('可训练参数量：{:,}'.format(parameter_number))

    best_accuracy = 0.0
    min_loss = float('inf')
    final_accuracy = None
    final_loss = None
    start_time = time.perf_counter()

    for epoch in range(EPOCHS):
        model.train()
        running_loss = 0.0

        for step, (inputs, labels) in enumerate(train_loader, start=0):
            optimizer.zero_grad()
            outputs = model(inputs)
            loss = loss_function(outputs, labels)
            loss.backward()
            optimizer.step()

            running_loss += loss.item()
            if step % LOG_INTERVAL == LOG_INTERVAL - 1:
                model.eval()
                with torch.no_grad():
                    outputs = model(val_image)
                    predictions = torch.max(outputs, dim=1)[1]
                    accuracy = (
                        torch.eq(predictions, val_label).sum().item()
                        / val_label.size(0)
                    )

                train_loss = running_loss / LOG_INTERVAL
                elapsed_time = time.perf_counter() - start_time
                best_accuracy = max(best_accuracy, accuracy)
                min_loss = min(min_loss, train_loss)
                final_accuracy = accuracy
                final_loss = train_loss

                print(
                    '[{}] [{:d}, {:5d}] train_loss: {:.3f}  '
                    'test_accuracy: {:.3f}  elapsed: {:.1f}s'.format(
                        config_name, epoch + 1, step + 1,
                        train_loss, accuracy, elapsed_time
                    )
                )

                append_log_row(
                    log_sheet, config_name, config, parameter_number,
                    epoch + 1, step + 1, train_loss, accuracy, elapsed_time
                )
                workbook.save(EXCEL_PATH)
                running_loss = 0.0
                model.train()

    training_time = time.perf_counter() - start_time
    if final_accuracy is None or final_loss is None:
        raise RuntimeError('训练数据不足{}个batch，未产生任何记录。'.format(
            LOG_INTERVAL
        ))

    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    model_path = MODELS_DIR / '{}.pth'.format(config_name)
    torch.save(model.state_dict(), model_path)

    baseline_metrics = read_baseline_metrics(summary_sheet)
    if config_name == 'baseline_current':
        baseline_metrics = {
            'parameters': parameter_number,
            'accuracy': final_accuracy,
            'loss': final_loss,
        }
    conclusion = build_conclusion(
        config_name, parameter_number, final_accuracy, baseline_metrics
    )

    append_summary_row(
        summary_sheet, config_name, config, parameter_number,
        final_accuracy, final_loss, best_accuracy, min_loss,
        training_time, model_path, conclusion
    )
    workbook.save(EXCEL_PATH)
    write_markdown_summary(summary_sheet, SUMMARY_PATH)

    print('实验完成：{}，模型保存至 {}'.format(config_name, model_path))


def create_data_loaders():
    transform = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5)),
    ])

    try:
        train_set = torchvision.datasets.CIFAR10(
            root=str(DATA_DIR), train=True,
            download=True, transform=transform
        )
        val_set = torchvision.datasets.CIFAR10(
            root=str(DATA_DIR), train=False,
            download=True, transform=transform
        )
    except (PermissionError, RuntimeError) as error:
        raise RuntimeError(
            '无法读取共享CIFAR-10目录：{}。请检查data目录访问权限和数据完整性。'
            .format(DATA_DIR)
        ) from error

    train_loader = torch.utils.data.DataLoader(
        train_set, batch_size=BATCH_SIZE, shuffle=True, num_workers=0
    )

    val_loader = torch.utils.data.DataLoader(
        val_set, batch_size=VALIDATION_BATCH_SIZE,
        shuffle=False, num_workers=0
    )
    val_image, val_label = next(iter(val_loader))
    return train_loader, val_image, val_label


def parse_args():
    parser = argparse.ArgumentParser(
        description='LeNet模型结构消融实验'
    )
    parser.add_argument(
        '--list', action='store_true', dest='list_only',
        help='列出全部实验配置后退出'
    )
    parser.add_argument(
        '--experiments', nargs='+', default=['all'],
        choices=['all'] + list(MODEL_CONFIGS.keys()),
        help='指定要运行的实验，默认依次运行全部实验'
    )
    return parser.parse_args()


def main():
    args = parse_args()
    if args.list_only:
        for name, config in MODEL_CONFIGS.items():
            print('{} | {} | {}'.format(
                name, config['category'], config['modification']
            ))
        return

    selected_experiments = (
        list(MODEL_CONFIGS.keys())
        if 'all' in args.experiments
        else args.experiments
    )

    try:
        workbook, log_sheet, summary_sheet = prepare_workbook(EXCEL_PATH)
    except PermissionError:
        print('无法写入结果文件：{}'.format(EXCEL_PATH))
        print('请关闭正在使用该文件的Excel或WPS，然后重新运行。')
        return

    try:
        train_loader, val_image, val_label = create_data_loaders()
        loss_function = nn.CrossEntropyLoss()

        for config_name in selected_experiments:
            run_experiment(
                config_name,
                train_loader,
                val_image,
                val_label,
                loss_function,
                workbook,
                log_sheet,
                summary_sheet,
            )
    finally:
        workbook.save(EXCEL_PATH)
        workbook.close()

    print('\n所选结构消融实验已完成。')
    print('Excel结果：{}'.format(EXCEL_PATH))
    print('Markdown汇总：{}'.format(SUMMARY_PATH))


if __name__ == '__main__':
    main()
