"""LeNet 模型参数量与测试准确率效率分析。

这是纯后处理脚本：读取已有 xlsx/csv/json 结果，不训练模型，也不修改源数据。
"""

import argparse
import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use('Agg')

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.lines import Line2D
from matplotlib.ticker import FuncFormatter, PercentFormatter
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
EXPERIMENT_DIR = SCRIPT_DIR.parent
DEFAULT_INPUT_CANDIDATES = [
    EXPERIMENT_DIR / 'results' / 'model_structure_results.xlsx',
    EXPERIMENT_DIR / 'results' / 'model_structure_results.csv',
    EXPERIMENT_DIR / 'results' / 'model_structure_results.json',
]

CSV_FILENAME = 'parameter_accuracy_results.csv'
REPORT_FILENAME = 'parameter_efficiency_report.md'
TRADEOFF_FILENAME = 'parameter_accuracy_tradeoff.png'
BUBBLE_FILENAME = 'parameter_efficiency_bubble.png'

FIELD_ALIASES = {
    'model_name': [
        'model_name', 'model name', 'model', 'experiment', 'name', '配置名'
    ],
    'parameters': [
        'parameters', 'parameter', 'parameter count', 'parameter number',
        'params', '可训练参数量', '参数量'
    ],
    # 优先读取最佳准确率；缺少时再回退到最终/普通准确率字段。
    'accuracy': [
        'best accuracy', 'best test accuracy', '最佳准确率',
        'test accuracy', 'accuracy', 'final accuracy', '测试准确率'
    ],
    'training_time': [
        'training time (s)', 'training time', 'training_time',
        'elapsed time (s)', 'time', '训练时间'
    ],
    'category': [
        'category', 'experiment category', '实验类别', '类别'
    ],
    'modification': [
        'modification', 'change', '修改', '结构修改'
    ],
}

DISPLAY_NAMES = {
    'baseline_current': 'Baseline',
    'channel_reduced_4_8': 'Channel-4-8',
    'channel_classic_6_16': 'Channel-6-16',
    'depth_add_conv': 'Add-Conv3',
    'kernel_size_3': 'Kernel-3',
    'kernel_size_7': 'Kernel-7',
    'pooling_avg': 'AvgPool',
    'pooling_none': 'No-Pooling',
    'activation_sigmoid': 'Sigmoid',
    'activation_tanh': 'Tanh',
    'activation_leaky_relu': 'LeakyReLU',
    'batch_norm': 'BatchNorm',
    'dropout_0_5': 'Dropout-0.5',
}

CATEGORY_LABELS = {
    '基线': 'Baseline',
    '卷积通道消融': 'Channel',
    '网络深度消融': 'Depth',
    '卷积核大小消融': 'Kernel',
    '池化方式消融': 'Pooling',
    '激活函数消融': 'Activation',
    'Batch Normalization消融': 'BatchNorm',
    'Dropout消融': 'Dropout',
}

CATEGORY_MARKERS = {
    '基线': '*',
    '卷积通道消融': 'o',
    '网络深度消融': 's',
    '卷积核大小消融': '^',
    '池化方式消融': 'D',
    '激活函数消融': 'P',
    'Batch Normalization消融': 'X',
    'Dropout消融': 'v',
}

CATEGORY_COLORS = {
    '基线': '#D62728',
    '卷积通道消融': '#1F77B4',
    '网络深度消融': '#9467BD',
    '卷积核大小消融': '#FF7F0E',
    '池化方式消融': '#2CA02C',
    '激活函数消融': '#17BECF',
    'Batch Normalization消融': '#E377C2',
    'Dropout消融': '#8C564B',
}

ANNOTATION_OFFSETS = {
    'baseline_current': (10, 5),
    'batch_norm': (7, 8),
    'activation_leaky_relu': (9, -25),
    'pooling_avg': (-10, -18),
    'kernel_size_3': (-7, -18),
    'depth_add_conv': (-7, 9),
    'channel_classic_6_16': (7, -17),
    'channel_reduced_4_8': (7, 8),
    'kernel_size_7': (7, 8),
    'pooling_none': (-7, -18),
    'activation_sigmoid': (7, -18),
    'activation_tanh': (7, 8),
    'dropout_0_5': (-7, -18),
}

BUBBLE_ANNOTATION_OFFSETS = {
    **ANNOTATION_OFFSETS,
    'baseline_current': (15, 2),
    'batch_norm': (12, 12),
    'activation_leaky_relu': (14, -28),
    'pooling_avg': (-14, 12),
    'kernel_size_3': (14, -38),
    'activation_tanh': (14, 12),
    'activation_sigmoid': (14, -22),
    'dropout_0_5': (-14, -26),
    'depth_add_conv': (-12, 14),
}


@dataclass(frozen=True)
class ModelRecord:
    model_name: str
    parameters: int
    accuracy: float  # 0~1
    training_time: float | None
    category: str
    modification: str


def normalize_key(value):
    return ''.join(
        character for character in str(value).strip().lower()
        if character.isalnum()
    )


def parse_number(value, field_name, required=True):
    if value is None or str(value).strip() == '':
        if required:
            raise ValueError('字段{}缺少数值。'.format(field_name))
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(',', '')
    is_percent = text.endswith('%')
    text = text.rstrip('%').rstrip('s').strip()
    number = float(text)
    return number / 100 if is_percent else number


def get_field(raw_record, field_name, required=True, default=None):
    normalized_record = {
        normalize_key(key): value for key, value in raw_record.items()
    }
    for alias in FIELD_ALIASES[field_name]:
        normalized_alias = normalize_key(alias)
        if normalized_alias in normalized_record:
            value = normalized_record[normalized_alias]
            if value is not None and str(value).strip() != '':
                return value
    if required:
        raise ValueError(
            '记录缺少字段{}，支持别名：{}'.format(
                field_name, FIELD_ALIASES[field_name]
            )
        )
    return default


def standardize_record(raw_record):
    model_name = str(get_field(raw_record, 'model_name')).strip()
    parameters = int(round(parse_number(
        get_field(raw_record, 'parameters'), 'parameters'
    )))
    accuracy = parse_number(
        get_field(raw_record, 'accuracy'), 'accuracy'
    )
    if accuracy > 1:
        accuracy /= 100
    training_time = parse_number(
        get_field(raw_record, 'training_time', required=False),
        'training_time', required=False
    )
    category = str(get_field(
        raw_record, 'category', required=False, default='未分类'
    )).strip()
    modification = str(get_field(
        raw_record, 'modification', required=False, default=''
    )).strip()

    if parameters <= 0:
        raise ValueError('{}的参数量必须大于0。'.format(model_name))
    if not 0 <= accuracy <= 1:
        raise ValueError('{}的准确率必须位于0~1或0~100%。'.format(model_name))
    if training_time is not None and training_time < 0:
        raise ValueError('{}的训练时间不能为负数。'.format(model_name))

    return ModelRecord(
        model_name=model_name,
        parameters=parameters,
        accuracy=accuracy,
        training_time=training_time,
        category=category,
        modification=modification,
    )


def load_xlsx(input_path):
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = (
            workbook['summary']
            if 'summary' in workbook.sheetnames
            else workbook.active
        )
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows))
        return [
            dict(zip(headers, values))
            for values in rows
            if any(value is not None for value in values)
        ]
    finally:
        workbook.close()


def load_csv(input_path):
    with input_path.open('r', encoding='utf-8-sig', newline='') as file:
        return list(csv.DictReader(file))


def load_json(input_path):
    with input_path.open('r', encoding='utf-8') as file:
        payload = json.load(file)

    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ('summary', 'records', 'models', 'data', 'results'):
            if isinstance(payload.get(key), list):
                return payload[key]
        if all(isinstance(value, dict) for value in payload.values()):
            return [
                {'model_name': name, **values}
                for name, values in payload.items()
            ]
    raise ValueError('JSON应为记录列表，或包含summary/records/models/data列表。')


def find_default_input():
    for candidate in DEFAULT_INPUT_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        '未找到默认结果文件，请通过--input指定xlsx/csv/json。'
    )


def load_records(input_path):
    suffix = input_path.suffix.lower()
    loaders = {
        '.xlsx': load_xlsx,
        '.csv': load_csv,
        '.json': load_json,
    }
    if suffix not in loaders:
        raise ValueError('不支持的输入格式：{}。'.format(suffix))

    raw_records = loaders[suffix](input_path)
    records_by_name = {}
    for raw_record in raw_records:
        record = standardize_record(raw_record)
        records_by_name[record.model_name] = record
    records = list(records_by_name.values())
    if len(records) < 2:
        raise ValueError('参数效率分析至少需要两条模型记录。')
    return records


def find_baseline(records, requested_name=None):
    if requested_name:
        for record in records:
            if record.model_name == requested_name:
                return record
        raise ValueError('未找到指定基线：{}'.format(requested_name))

    for record in records:
        if record.model_name == 'baseline_current':
            return record
    for record in records:
        if '基线' in record.category or 'baseline' in record.model_name.lower():
            return record
    raise ValueError('无法自动识别Baseline，请使用--baseline指定模型名。')


def mark_pareto_front(records):
    pareto_names = set()
    for record in records:
        dominated = any(
            other.model_name != record.model_name
            and other.parameters <= record.parameters
            and other.accuracy >= record.accuracy
            and (
                other.parameters < record.parameters
                or other.accuracy > record.accuracy
            )
            for other in records
        )
        if not dominated:
            pareto_names.add(record.model_name)
    return pareto_names


def analyze_records(records, baseline):
    pareto_names = mark_pareto_front(records)
    analyzed = []
    for record in records:
        parameter_growth_ratio = (
            record.parameters - baseline.parameters
        ) / baseline.parameters
        accuracy_change = record.accuracy - baseline.accuracy
        accuracy_growth_ratio = accuracy_change / baseline.accuracy
        efficiency_ratio = (
            accuracy_growth_ratio / parameter_growth_ratio
            if not math.isclose(parameter_growth_ratio, 0.0, abs_tol=1e-12)
            else None
        )
        analyzed.append({
            'record': record,
            'parameter_growth_ratio': parameter_growth_ratio,
            'accuracy_change_pp': accuracy_change * 100,
            'accuracy_growth_ratio': accuracy_growth_ratio,
            'efficiency_ratio': efficiency_ratio,
            'accuracy_per_100k_parameters': (
                record.accuracy * 100 / (record.parameters / 100000)
            ),
            'pareto_efficient': record.model_name in pareto_names,
        })
    return analyzed


def configure_plot_style():
    plt.rcParams.update({
        'font.sans-serif': [
            'Microsoft YaHei', 'SimHei', 'Noto Sans CJK SC', 'DejaVu Sans'
        ],
        'axes.unicode_minus': False,
        'axes.spines.top': False,
        'axes.spines.right': False,
        'axes.titleweight': 'bold',
        'axes.grid': True,
        'grid.linestyle': '--',
        'grid.alpha': 0.28,
        'figure.facecolor': 'white',
        'axes.facecolor': 'white',
    })


def category_styles(records):
    fallback_markers = ['o', 's', '^', 'D', 'P', 'X', 'v', '<', '>']
    fallback_colors = plt.get_cmap('tab10').colors
    categories = list(dict.fromkeys(record.category for record in records))
    styles = {}
    for index, category in enumerate(categories):
        styles[category] = {
            'marker': CATEGORY_MARKERS.get(
                category, fallback_markers[index % len(fallback_markers)]
            ),
            'color': CATEGORY_COLORS.get(
                category, fallback_colors[index % len(fallback_colors)]
            ),
            'label': CATEGORY_LABELS.get(category, category),
        }
    return styles


def parameter_tick(value, _position):
    if value >= 1000000:
        return '{:g}M'.format(value / 1000000)
    if value >= 1000:
        return '{:g}K'.format(value / 1000)
    return '{:g}'.format(value)


def annotate_points(axis, records, offsets=None):
    offsets = ANNOTATION_OFFSETS if offsets is None else offsets
    for record in records:
        x_offset, y_offset = offsets.get(
            record.model_name, (6, 6)
        )
        axis.annotate(
            DISPLAY_NAMES.get(record.model_name, record.model_name),
            (record.parameters, record.accuracy * 100),
            xytext=(x_offset, y_offset),
            textcoords='offset points',
            ha='right' if x_offset < 0 else 'left',
            va='top' if y_offset < 0 else 'bottom',
            fontsize=8.5,
            bbox={
                'boxstyle': 'round,pad=0.18',
                'facecolor': 'white',
                'edgecolor': 'none',
                'alpha': 0.68,
            },
        )


def add_baseline_guides(axis, baseline):
    axis.axvline(
        baseline.parameters, color='#D62728', linestyle=':',
        linewidth=1.1, alpha=0.65
    )
    axis.axhline(
        baseline.accuracy * 100, color='#D62728', linestyle=':',
        linewidth=1.1, alpha=0.65
    )


def plot_parameter_accuracy_tradeoff(records, baseline, output_path):
    configure_plot_style()
    styles = category_styles(records)
    figure, axis = plt.subplots(figsize=(12.8, 7.4))

    for category, style in styles.items():
        category_records = [
            record for record in records if record.category == category
        ]
        sizes = [230 if record.model_name == baseline.model_name else 95
                 for record in category_records]
        axis.scatter(
            [record.parameters for record in category_records],
            [record.accuracy * 100 for record in category_records],
            s=sizes,
            marker=style['marker'],
            c=[style['color']],
            edgecolors='black' if category == baseline.category else 'white',
            linewidths=1.3 if category == baseline.category else 0.8,
            alpha=0.92,
            label=style['label'],
            zorder=4 if category == baseline.category else 3,
        )

    pareto_records = sorted(
        [record for record in records
         if record.model_name in mark_pareto_front(records)],
        key=lambda record: record.parameters
    )
    axis.plot(
        [record.parameters for record in pareto_records],
        [record.accuracy * 100 for record in pareto_records],
        color='#555555', linestyle='--', linewidth=1.25,
        alpha=0.7, label='Pareto frontier', zorder=2
    )

    annotate_points(axis, records)
    add_baseline_guides(axis, baseline)
    axis.set_xscale('log')
    axis.xaxis.set_major_formatter(FuncFormatter(parameter_tick))
    axis.set_xlabel('Trainable Parameter Count (log scale)')
    axis.set_ylabel('Best Test Accuracy (%)')
    axis.set_title('LeNet Parameter–Accuracy Trade-off')
    accuracies = [record.accuracy * 100 for record in records]
    axis.set_ylim(max(0, min(accuracies) - 4), min(100, max(accuracies) + 4))
    axis.legend(
        title='Experiment Category', loc='center left',
        bbox_to_anchor=(1.01, 0.5), frameon=False
    )
    figure.tight_layout()
    figure.savefig(output_path, dpi=240, bbox_inches='tight')
    plt.close(figure)


def bubble_area(training_time, minimum, maximum, fallback):
    value = fallback if training_time is None else training_time
    if math.isclose(minimum, maximum):
        return 750
    return 320 + 1150 * (value - minimum) / (maximum - minimum)


def plot_parameter_efficiency_bubble(records, baseline, output_path):
    configure_plot_style()
    styles = category_styles(records)
    valid_times = [
        record.training_time for record in records
        if record.training_time is not None
    ]
    fallback_time = float(np.median(valid_times)) if valid_times else 1.0
    minimum_time = min(valid_times) if valid_times else fallback_time
    maximum_time = max(valid_times) if valid_times else fallback_time

    figure, axis = plt.subplots(figsize=(13.2, 7.6))
    for category, style in styles.items():
        category_records = [
            record for record in records if record.category == category
        ]
        axis.scatter(
            [record.parameters for record in category_records],
            [record.accuracy * 100 for record in category_records],
            s=[bubble_area(
                record.training_time, minimum_time,
                maximum_time, fallback_time
            ) for record in category_records],
            marker=style['marker'],
            c=[style['color']],
            edgecolors='black' if category == baseline.category else 'white',
            linewidths=1.5 if category == baseline.category else 1.0,
            alpha=0.62,
            zorder=4 if category == baseline.category else 3,
        )

    annotate_points(axis, records, BUBBLE_ANNOTATION_OFFSETS)
    add_baseline_guides(axis, baseline)
    axis.set_xscale('log')
    axis.xaxis.set_major_formatter(FuncFormatter(parameter_tick))
    axis.set_xlabel('Trainable Parameter Count (log scale)')
    axis.set_ylabel('Best Test Accuracy (%)')
    axis.set_title('LeNet Parameter Efficiency Bubble Chart')
    accuracies = [record.accuracy * 100 for record in records]
    axis.set_ylim(max(0, min(accuracies) - 4), min(100, max(accuracies) + 4))

    category_handles = [
        Line2D(
            [0], [0], marker=style['marker'], color='none',
            markerfacecolor=style['color'], markeredgecolor='white',
            markersize=9, label=style['label']
        )
        for style in styles.values()
    ]
    category_legend = axis.legend(
        handles=category_handles, title='Experiment Category',
        loc='upper left', bbox_to_anchor=(1.01, 1.0), frameon=False
    )
    axis.add_artist(category_legend)

    time_legend = None
    if valid_times:
        time_levels = sorted(set([
            round(minimum_time), round(fallback_time), round(maximum_time)
        ]))
        time_handles = [
            axis.scatter(
                [], [],
                s=bubble_area(value, minimum_time, maximum_time, fallback_time),
                facecolors='none', edgecolors='#666666', linewidths=1,
                label='{} s'.format(value)
            )
            for value in time_levels
        ]
        time_legend = axis.legend(
            handles=time_handles, title='Training Time',
            loc='lower left', bbox_to_anchor=(1.01, 0.0),
            frameon=False, labelspacing=1.6
        )

    figure.tight_layout(rect=(0, 0, 0.78, 1))
    extra_artists = (
        (category_legend, time_legend)
        if time_legend is not None else (category_legend,)
    )
    figure.savefig(
        output_path, dpi=240, bbox_inches='tight',
        bbox_extra_artists=extra_artists
    )
    plt.close(figure)


def write_results_csv(analyzed, baseline, output_path):
    fieldnames = [
        'model_name', 'parameters', 'accuracy', 'training_time', 'category',
        'modification', 'is_baseline', 'parameter_growth_ratio',
        'parameter_growth_percent', 'accuracy_change_pp',
        'accuracy_growth_ratio', 'efficiency_ratio',
        'accuracy_per_100k_parameters', 'pareto_efficient'
    ]
    with output_path.open('w', encoding='utf-8-sig', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for item in analyzed:
            record = item['record']
            writer.writerow({
                'model_name': record.model_name,
                'parameters': record.parameters,
                'accuracy': round(record.accuracy * 100, 4),
                'training_time': (
                    round(record.training_time, 4)
                    if record.training_time is not None else ''
                ),
                'category': record.category,
                'modification': record.modification,
                'is_baseline': record.model_name == baseline.model_name,
                'parameter_growth_ratio': round(
                    item['parameter_growth_ratio'], 8
                ),
                'parameter_growth_percent': round(
                    item['parameter_growth_ratio'] * 100, 4
                ),
                'accuracy_change_pp': round(item['accuracy_change_pp'], 4),
                'accuracy_growth_ratio': round(
                    item['accuracy_growth_ratio'], 8
                ),
                'efficiency_ratio': (
                    round(item['efficiency_ratio'], 8)
                    if item['efficiency_ratio'] is not None else ''
                ),
                'accuracy_per_100k_parameters': round(
                    item['accuracy_per_100k_parameters'], 4
                ),
                'pareto_efficient': item['pareto_efficient'],
            })


def format_time(value):
    return '{:.1f} s'.format(value) if value is not None else 'N/A'


def format_efficiency(value):
    return '{:.3f}'.format(value) if value is not None else 'N/A'


def find_item(analyzed, model_name):
    return next(
        (item for item in analyzed
         if item['record'].model_name == model_name),
        None
    )


def write_report(analyzed, baseline, input_path, output_path):
    records = [item['record'] for item in analyzed]
    best_item = max(analyzed, key=lambda item: item['record'].accuracy)
    improvement_candidates = [
        item for item in analyzed
        if item['parameter_growth_ratio'] > 0
        and item['accuracy_change_pp'] > 0
        and item['efficiency_ratio'] is not None
    ]
    best_efficiency_item = (
        max(improvement_candidates, key=lambda item: item['efficiency_ratio'])
        if improvement_candidates else None
    )
    pareto_items = sorted(
        [item for item in analyzed if item['pareto_efficient']],
        key=lambda item: item['record'].parameters
    )

    log_parameter_correlation = float(np.corrcoef(
        np.log10([record.parameters for record in records]),
        [record.accuracy * 100 for record in records]
    )[0, 1])

    lines = [
        '# LeNet模型参数量—测试准确率效率分析报告',
        '',
        '## 1. 实验目的',
        '',
        '本分析基于已完成的LeNet结构消融实验，研究可训练参数量、最佳测试准确率和训练时间之间的关系。分析不重新训练模型，也不修改原始结果。数据源为 `{}`，Accuracy统一读取“最佳测试准确率”。'.format(input_path),
        '',
        '以 `{}` 为基线，参数增长比例、Accuracy变化和参数效率分别定义为：'.format(baseline.model_name),
        '',
        '- 参数增长比例：`(parameters_new - parameters_baseline) / parameters_baseline`。',
        '- Accuracy变化：`accuracy_new - accuracy_baseline`，报告中以百分点表示。',
        '- Accuracy提升比例：`(accuracy_new - accuracy_baseline) / accuracy_baseline`。',
        '- 参数效率：`Accuracy提升比例 / 参数增长比例`。参数量不变时该值记为N/A；减参模型的正值表示“参数减少比例大于准确率损失比例”，不等同于准确率提升。',
        '',
        '## 2. 数据统计表',
        '',
        '| Model | Category | Parameters | Best Accuracy | Training Time | Pareto |',
        '| --- | --- | ---: | ---: | ---: | :---: |',
    ]
    for item in analyzed:
        record = item['record']
        lines.append(
            '| {} | {} | {:,} | {:.2%} | {} | {} |'.format(
                DISPLAY_NAMES.get(record.model_name, record.model_name),
                CATEGORY_LABELS.get(record.category, record.category),
                record.parameters,
                record.accuracy,
                format_time(record.training_time),
                '✓' if item['pareto_efficient'] else '',
            )
        )

    lines.extend([
        '',
        '全部模型的 `log10(参数量)` 与最佳准确率的Pearson相关系数为 **{:.3f}**。这反映当前13个配置的整体关联，但不能作为参数量导致准确率变化的因果证据。'.format(log_parameter_correlation),
        '',
        '![Parameter–Accuracy Trade-off]({})'.format(TRADEOFF_FILENAME),
        '',
        '![Parameter Efficiency Bubble Chart]({})'.format(BUBBLE_FILENAME),
        '',
        '## 3. 最佳性能模型',
        '',
        '**{}** 获得最高最佳测试准确率 **{:.2%}**，参数量为 **{:,}**，训练时间为 **{}**。相较Baseline，其准确率变化为 **{:+.2f}个百分点**，参数量变化为 **{:+.2%}**。'.format(
            DISPLAY_NAMES.get(
                best_item['record'].model_name,
                best_item['record'].model_name
            ),
            best_item['record'].accuracy,
            best_item['record'].parameters,
            format_time(best_item['record'].training_time),
            best_item['accuracy_change_pp'],
            best_item['parameter_growth_ratio'],
        ),
        '',
        '该模型以很小的参数增量获得明确性能提升，说明结构设计和优化稳定性比单纯扩大模型规模更重要。',
        '',
        '## 4. 参数效率最高模型',
        '',
        '| Model | Parameter Growth | Accuracy Change | Accuracy Growth | Efficiency |',
        '| --- | ---: | ---: | ---: | ---: |',
    ])
    for item in analyzed:
        record = item['record']
        lines.append(
            '| {} | {:+.2%} | {:+.2f} pp | {:+.2%} | {} |'.format(
                DISPLAY_NAMES.get(record.model_name, record.model_name),
                item['parameter_growth_ratio'],
                item['accuracy_change_pp'],
                item['accuracy_growth_ratio'],
                format_efficiency(item['efficiency_ratio']),
            )
        )

    if best_efficiency_item is not None:
        efficient_record = best_efficiency_item['record']
        lines.extend([
            '',
            '在“参数增加且准确率提高”的模型中，**{}** 的题设参数效率最高，效率比值为 **{:.3f}**。其参数只增加 **{:+.2%}**，Accuracy提高 **{:+.2f}个百分点**。'.format(
                DISPLAY_NAMES.get(
                    efficient_record.model_name,
                    efficient_record.model_name
                ),
                best_efficiency_item['efficiency_ratio'],
                best_efficiency_item['parameter_growth_ratio'],
                best_efficiency_item['accuracy_change_pp'],
            ),
        ])

    lines.extend([
        '',
        'Pareto前沿模型为：{}。这些模型不存在“参数更少且准确率不低”的其他配置，代表不同模型规模下的有效折中点。'.format(
            '、'.join(
                DISPLAY_NAMES.get(
                    item['record'].model_name,
                    item['record'].model_name
                )
                for item in pareto_items
            )
        ),
        '',
        '## 5. 结构分析',
        '',
    ])

    by_name = {item['record'].model_name: item for item in analyzed}
    batch_norm = by_name.get('batch_norm')
    dropout = by_name.get('dropout_0_5')
    kernel3 = by_name.get('kernel_size_3')
    kernel7 = by_name.get('kernel_size_7')
    channel4 = by_name.get('channel_reduced_4_8')
    channel6 = by_name.get('channel_classic_6_16')
    depth = by_name.get('depth_add_conv')
    no_pooling = by_name.get('pooling_none')

    if batch_norm:
        lines.append(
            '- **BatchNorm：** 仅增加{:,}个参数（{:+.2%}），最佳准确率提高{:+.2f}个百分点，是最典型的高参数效率修改。'.format(
                batch_norm['record'].parameters - baseline.parameters,
                batch_norm['parameter_growth_ratio'],
                batch_norm['accuracy_change_pp'],
            )
        )
    if dropout:
        lines.append(
            '- **Dropout：** 参数量不变，但最佳准确率变化{:+.2f}个百分点；在固定10轮训练和p=0.5下，正则化收益有限。'.format(
                dropout['accuracy_change_pp']
            )
        )
    if kernel3 and kernel7:
        lines.append(
            '- **卷积核：** Kernel-3增加{:+.2%}参数但准确率变化{:+.2f}个百分点；Kernel-7因无padding卷积使展平维度缩小，参数反而减少{:.2%}，准确率下降{:.2f}个百分点。参数变化主要来自后续FC输入尺寸，而不只是卷积核本身。'.format(
                kernel3['parameter_growth_ratio'],
                kernel3['accuracy_change_pp'],
                abs(kernel7['parameter_growth_ratio']),
                abs(kernel7['accuracy_change_pp']),
            )
        )
    if channel4 and channel6:
        lines.append(
            '- **通道数量：** Channel-4-8和Channel-6-16分别减少{:.2%}和{:.2%}参数，同时损失{:.2f}和{:.2f}个百分点。通道增加能提升容量，但轻量化配置提供了可选择的精度—规模折中。'.format(
                abs(channel4['parameter_growth_ratio']),
                abs(channel6['parameter_growth_ratio']),
                abs(channel4['accuracy_change_pp']),
                abs(channel6['accuracy_change_pp']),
            )
        )
    if depth:
        lines.append(
            '- **增加Conv3：** 参数减少{:.2%}而最佳准确率只下降{:.2f}个百分点，位于Pareto前沿。额外池化缩小FC输入，使其成为比Channel-6-16更有效的紧凑配置，但本结果同时包含“增深”和“进一步下采样”的共同影响。'.format(
                abs(depth['parameter_growth_ratio']),
                abs(depth['accuracy_change_pp']),
            )
        )
    if no_pooling:
        lines.append(
            '- **取消Pooling：** 参数增长{:+.2%}、训练时间达到{}，但最佳准确率变化{:+.2f}个百分点，属于显著增加计算成本却收益为负的结构。'.format(
                no_pooling['parameter_growth_ratio'],
                format_time(no_pooling['record'].training_time),
                no_pooling['accuracy_change_pp'],
            )
        )

    lines.extend([
        '',
        '气泡图中，理想模型位于左上方且气泡较小，即参数少、准确率高、训练时间短。Add-Conv3在紧凑模型中表现突出；BatchNorm位于Baseline附近但准确率更高；No-Pooling位于最右侧且气泡最大，计算成本最高但性能没有改善。',
        '',
        '## 6. 总结',
        '',
        '实验结果否定了“增加参数一定提升性能”的假设。No-Pooling将参数量提高到Baseline的约18.5倍，却降低测试准确率；Kernel-3也在增加参数后未获得准确率提升。相反，BatchNorm几乎不改变参数规模便取得全组最高准确率，说明有效的特征归一化和优化稳定性比盲目扩容更重要。',
        '',
        '从trade-off角度看：若追求最高性能，BatchNorm是最佳选择；若追求紧凑模型，Add-Conv3以约41%的Baseline参数保留了接近Baseline的准确率；若需要更小规模，可进一步选择Channel-4-8或Channel-6-16并接受相应精度损失。',
        '',
        '> 结论仅适用于当前CIFAR-10测试方法、单一随机种子、Adam、学习率0.001和10轮训练设置。不同结构可能具有不同的最优超参数，训练时间也会受硬件与系统负载影响。',
        '',
    ])
    output_path.write_text('\n'.join(lines), encoding='utf-8')


def parse_args():
    parser = argparse.ArgumentParser(
        description='LeNet模型参数量—测试准确率效率分析'
    )
    parser.add_argument(
        '--input', type=Path, default=None,
        help='已有实验结果文件，支持xlsx/csv/json；默认自动发现'
    )
    parser.add_argument(
        '--baseline', default=None,
        help='基线模型名称；默认自动识别baseline_current或基线类别'
    )
    parser.add_argument(
        '--output-dir', type=Path, default=SCRIPT_DIR,
        help='CSV、报告和图片输出目录，默认当前analysis目录'
    )
    return parser.parse_args()


def main():
    args = parse_args()
    input_path = (
        args.input.resolve() if args.input is not None
        else find_default_input().resolve()
    )
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    records = load_records(input_path)
    baseline = find_baseline(records, args.baseline)
    analyzed = analyze_records(records, baseline)

    csv_path = output_dir / CSV_FILENAME
    report_path = output_dir / REPORT_FILENAME
    tradeoff_path = output_dir / TRADEOFF_FILENAME
    bubble_path = output_dir / BUBBLE_FILENAME

    write_results_csv(analyzed, baseline, csv_path)
    plot_parameter_accuracy_tradeoff(records, baseline, tradeoff_path)
    plot_parameter_efficiency_bubble(records, baseline, bubble_path)
    write_report(analyzed, baseline, input_path, report_path)

    print('参数效率分析完成：')
    print('输入记录数：{}'.format(len(records)))
    print('Baseline：{}'.format(baseline.model_name))
    for output_path in (csv_path, report_path, tradeoff_path, bubble_path):
        print(output_path)


if __name__ == '__main__':
    main()
